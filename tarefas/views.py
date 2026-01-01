"""
VIEWS COMPLETAS - SISTEMA DE 2 NÍVEIS DE CRITICIDADE
Arquitetura Limpa e Modular - VERSÃO REFATORADA v4.0 (Binário)

✅ REFATORAÇÃO v4.0 - SIMPLIFICAÇÃO 2 NÍVEIS (CRÍTICA/REGULAR):
- KPIs e Gráficos atualizados para 2 níveis
- Filtros e ordenações ajustados
- Legendas e textos simplificados

✅ SUGESTÃO 1 APLICADA:
- Otimização N+1 na view 'lista_servidores' usando 'annotate'
  para calcular 'total', 'criticas' e 'regulares' em uma única query.

Arquivo: tarefas/views.py
Data: 27/10/2025
"""

from datetime import date, timedelta
import json
import sys
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import get_user_model
from django.db.models import Q, Count, Case, When, IntegerField, Prefetch, Sum
from django.core.paginator import Paginator
from django.http import JsonResponse, HttpResponse
from tarefas.models import Tarefa
from tarefas.parametros import ParametrosAnalise
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

User = get_user_model()


# ============================================
# HELPER: VERIFICAÇÃO DE PERMISSÕES
# ============================================
def usuario_eh_coordenador(user):
    """Verifica se o usuário é coordenador"""
    return user.groups.filter(name='Coordenador').exists() or user.is_staff


# ============================================
# HELPER: INFORMAÇÕES DE REGRAS
# ============================================
def get_regras_info():
    """
    Retorna informações detalhadas sobre todas as regras de criticidade.
    Usado para exibir cards informativos nos dashboards.

    IMPORTANTE: As chaves devem corresponder às constantes do AnalisadorCriticidade
    """
    # Buscar parâmetros atuais
    try:
        params = ParametrosAnalise.get_configuracao_ativa()
    except:
        params = None

    return {
        'REGRA_1_EXIGENCIA_CUMPRIDA': {
            'numero': 'REGRA 1',
            'nome': 'Exigência Cumprida - Aguardando Análise',
            'descricao': 'Servidor cadastrou exigência que foi cumprida pelo segurado. Aguardando análise dos documentos.',
            'prazo': f"{params.prazo_analise_exigencia_cumprida if params else 7} dias após cumprimento",
            'badge_class': 'bg-info',
            'icon': 'fas fa-file-check',
            'exemplo': 'Status: Pendente + Exigência cumprida (após atribuição)'
        },
        'REGRA_2_CUMPRIMENTO_EXIGENCIA': {
            'numero': 'REGRA 2',
            'nome': 'Cumprimento de Exigência pelo Segurado',
            'descricao': 'Segurado tem prazo para cumprir exigência. Após vencimento, servidor tem prazo para conclusão.',
            'prazo': f"Prazo + {params.prazo_tolerancia_exigencia if params else 5} dias (tolerância) + {params.prazo_servidor_apos_vencimento if params else 7} dias (conclusão)",
            'badge_class': 'bg-warning',
            'icon': 'fas fa-hourglass-half',
            'exemplo': 'Status: Cumprimento de exigência + Prazo definido'
        },
        'REGRA_3_PRIMEIRA_ACAO_SEM_EXIGENCIA': {
            'numero': 'REGRA 3',
            'nome': 'Tarefa Nunca Trabalhada',
            'descricao': 'Servidor puxou tarefa nova que nunca entrou em exigência. Tem prazo para primeira ação.',
            'prazo': f"{params.prazo_primeira_acao if params else 10} dias após puxar tarefa",
            'badge_class': 'bg-primary',
            'icon': 'fas fa-play-circle',
            'exemplo': 'Status: Pendente + Nunca entrou em exigência'
        },
        'REGRA_4_PRIMEIRA_ACAO_COM_EXIGENCIA': {
            'numero': 'REGRA 4',
            'nome': 'Exigência Cumprida Anterior',
            'descricao': 'Servidor puxou tarefa com exigência já cumprida (por outro servidor). Tem prazo para análise.',
            'prazo': f"{params.prazo_primeira_acao if params else 10} dias após atribuição",
            'badge_class': 'bg-success',
            'icon': 'fas fa-history',
            'exemplo': 'Status: Pendente + Exigência cumprida (antes da atribuição)'
        }
    }


def get_regras_resumo_servidor(tarefas_servidor):
    """
    Calcula resumo de tarefas por regra para um servidor específico.
    Usado para exibir cards de regras nos detalhes do servidor.
    """
    regras_info = get_regras_info()
    resumo = {}
    
    for regra_key, info in regras_info.items():
        count = tarefas_servidor.filter(regra_aplicada_calculado=regra_key).count()
        resumo[regra_key] = {
            'nome': info['nome'],
            'descricao': info['descricao'],
            'prazo': info['prazo'],
            'count': count,
            'badge_class': info['badge_class'],
            'icon': info['icon']
        }
    
    return resumo


# ============================================
# REDIRECIONAMENTO PÓS-LOGIN
# ============================================
@login_required
def redirect_after_login(request):
    """
    Redireciona o usuário após login baseado no grupo/perfil.

    - Coordenador → /tarefas/dashboard/coordenador/
    - Equipe Volante → /tarefas/equipe-volante/painel/
    - Servidor → /servidores/{siape}/
    - Outros → /admin/ (fallback seguro)
    """
    user = request.user

    try:
        # Verificar se é coordenador
        if user.groups.filter(name='Coordenador').exists():
            return redirect('tarefas:dashboard_coordenador')

        # Verificar se é da Equipe Volante
        elif user.groups.filter(name='Equipe Volante').exists():
            return redirect('tarefas:painel_equipe_volante')

        # Verificar se é servidor
        elif user.groups.filter(name='Servidor').exists():
            return redirect('tarefas:detalhe_servidor', siape=user.siape)

        # Outros usuários → redireciona para admin (fallback seguro)
        else:
            return redirect('/admin/')

    except Exception as e:
        # Se houver erro (ex: tabelas não existem), redireciona para admin
        # Isso permite que o sistema funcione mesmo sem migrations rodadas
        return redirect('/admin/')


# ============================================
# DASHBOARD COORDENADOR (SIMPLIFICADO)
# ============================================
@login_required
@user_passes_test(usuario_eh_coordenador)
def dashboard_coordenador(request):
    """
    Dashboard do coordenador - VERSÃO COM FILAS v5.0

    Exibe cards de todas as filas de trabalho com estatísticas.
    O coordenador pode clicar em cada card para ver detalhes da fila.

    Acesso: Apenas Coordenadores
    """
    from tarefas.filas import ORDEM_FILAS, obter_info_fila

    # Query otimizada: contar tarefas por fila (somente com responsável e ativas)
    stats_por_fila = Tarefa.objects.filter(
        siape_responsavel__isnull=False,
        ativa=True
    ).values('tipo_fila').annotate(
        total=Count('numero_protocolo_tarefa'),
        criticas=Count('numero_protocolo_tarefa', filter=Q(nivel_criticidade_calculado='CRÍTICA')),
        regulares=Count('numero_protocolo_tarefa', filter=Q(nivel_criticidade_calculado='REGULAR'))
    )

    # Transformar em dicionário para acesso rápido
    stats_dict = {item['tipo_fila']: item for item in stats_por_fila}

    # Construir lista ordenada de cards (apenas filas com tarefas)
    cards_filas = []
    for codigo_fila in ORDEM_FILAS:
        info = obter_info_fila(codigo_fila)
        stats = stats_dict.get(codigo_fila, {'total': 0, 'criticas': 0, 'regulares': 0})

        if stats['total'] > 0:  # Só mostrar filas com tarefas
            cards_filas.append({
                'codigo': codigo_fila,
                'nome': info['nome'],
                'nome_completo': info['nome_completo'],
                'descricao': info['descricao'],
                'cor': info['cor'],
                'cor_bootstrap': info['cor_bootstrap'],
                'icone': info['icone'],
                'total': stats['total'],
                'criticas': stats['criticas'],
                'regulares': stats['regulares'],
                'percentual_criticas': (stats['criticas'] / stats['total'] * 100) if stats['total'] > 0 else 0,
            })

    # Totais gerais
    total_geral = sum(c['total'] for c in cards_filas)
    criticas_geral = sum(c['criticas'] for c in cards_filas)
    regulares_geral = sum(c['regulares'] for c in cards_filas)

    context = {
        'cards_filas': cards_filas,
        'total_geral': total_geral,
        'criticas_geral': criticas_geral,
        'regulares_geral': regulares_geral,
        'percentual_criticas_geral': (criticas_geral / total_geral * 100) if total_geral > 0 else 0,
        'data_atualizacao': date.today(),
    }

    return render(request, 'dashboards/dashboard_coordenador.html', context)


# ============================================
# DETALHAMENTO DE FILA ESPECÍFICA
# ============================================
@login_required
def detalhe_fila(request, codigo_fila):
    """
    Detalhamento de uma fila de trabalho específica.

    Exibe:
    - Informações da fila (nome, descrição, cor)
    - Estatísticas (total, críticas, regulares)
    - Gráficos específicos da fila
    - Lista paginada de tarefas da fila

    Permissões:
    - Coordenador: Vê todas as tarefas da fila
    - Servidor: Vê apenas suas tarefas daquela fila
    """
    from tarefas.filas import obter_info_fila

    # Informações da fila
    info_fila = obter_info_fila(codigo_fila)

    # Filtro base: tarefas da fila com responsável e ativas
    tarefas = Tarefa.objects.filter(
        tipo_fila=codigo_fila,
        siape_responsavel__isnull=False,
        ativa=True
    ).select_related('siape_responsavel')

    # Se não for coordenador, mostrar apenas suas tarefas
    if not usuario_eh_coordenador(request.user):
        tarefas = tarefas.filter(siape_responsavel=request.user.siape)

    # Aplicar filtros do usuário (se houver)
    filtro_criticidade = request.GET.get('criticidade')
    if filtro_criticidade:
        tarefas = tarefas.filter(nivel_criticidade_calculado=filtro_criticidade)

    filtro_status = request.GET.get('status')
    if filtro_status:
        tarefas = tarefas.filter(status_tarefa=filtro_status)

    filtro_servidor = request.GET.get('servidor')
    if filtro_servidor and usuario_eh_coordenador(request.user):
        tarefas = tarefas.filter(siape_responsavel__siape=filtro_servidor)

    # Ordenação
    ordenacao = request.GET.get('ordem', '-pontuacao_criticidade')
    tarefas = tarefas.order_by(ordenacao)

    # OTIMIZAÇÃO: Calcular todas as estatísticas em uma única query usando aggregate
    from django.db.models import Sum, Case, When, IntegerField

    stats = tarefas.aggregate(
        total=Count('numero_protocolo_tarefa'),
        criticas=Count('numero_protocolo_tarefa', filter=Q(nivel_criticidade_calculado='CRÍTICA')),
        regulares=Count('numero_protocolo_tarefa', filter=Q(nivel_criticidade_calculado='REGULAR')),
        pendentes=Count('numero_protocolo_tarefa', filter=Q(status_tarefa='Pendente')),
        cumprimento=Count('numero_protocolo_tarefa', filter=Q(status_tarefa='Cumprimento de exigência')),
        outros=Count('numero_protocolo_tarefa', filter=~Q(status_tarefa__in=['Pendente', 'Cumprimento de exigência']))
    )

    # Extrair valores das estatísticas
    total = stats['total']
    criticas = stats['criticas']
    regulares = stats['regulares']

    # Estatísticas por status
    status_counts = {
        'pendentes': stats['pendentes'],
        'cumprimento': stats['cumprimento'],
        'outros': stats['outros'],
    }

    # Gráficos
    grafico_criticidade = {
        'labels': json.dumps(['Críticas', 'Regulares']),
        'data': json.dumps([criticas, regulares]),
        'colors': json.dumps(['#dc3545', '#28a745'])
    }

    grafico_status = {
        'labels': json.dumps(['Pendente', 'Cumprimento', 'Outros']),
        'data': json.dumps([status_counts['pendentes'], status_counts['cumprimento'], status_counts['outros']]),
        'colors': json.dumps(['#ffc107', '#17a2b8', '#6c757d'])
    }

    # OTIMIZAÇÃO: Prefetch de justificativas e solicitações de ajuda para evitar N+1 queries
    from tarefas.models import Justificativa, SolicitacaoAjuda
    tarefas = tarefas.prefetch_related(
        Prefetch(
            'justificativas',
            queryset=Justificativa.objects.filter(status='APROVADA').order_by('-data_submissao'),
            to_attr='justificativas_aprovadas'
        ),
        Prefetch(
            'solicitacoes_ajuda',
            queryset=SolicitacaoAjuda.objects.filter(status__in=['PENDENTE', 'EM_ATENDIMENTO']).order_by('-data_solicitacao'),
            to_attr='solicitacoes_ativas'
        )
    )

    # Paginação - Desabilitada ao filtrar por servidor específico
    if filtro_servidor:
        # Sem paginação - exibir todas as tarefas do servidor
        tarefas_page = tarefas
    else:
        # Com paginação - quando vendo todos os servidores
        paginator = Paginator(tarefas, 50)
        page = request.GET.get('page')
        tarefas_page = paginator.get_page(page)

    # Lista de servidores para filtro (se coordenador)
    servidores_fila = []
    servidor_filtrado = None
    if usuario_eh_coordenador(request.user):
        # OTIMIZAÇÃO: Usar values_list para retornar apenas campos necessários e distinct
        servidores_fila = User.objects.filter(
            tarefas_sob_responsabilidade__tipo_fila=codigo_fila,
            tarefas_sob_responsabilidade__isnull=False,
            tarefas_sob_responsabilidade__ativa=True
        ).distinct().only('siape', 'nome_completo').order_by('nome_completo')

        # Buscar objeto do servidor se estiver filtrando
        if filtro_servidor:
            try:
                servidor_filtrado = User.objects.only('siape', 'nome_completo').get(siape=filtro_servidor)
            except User.DoesNotExist:
                pass

    # Ranking dos 20 servidores com mais tarefas nesta fila
    ranking_servidores = []
    if usuario_eh_coordenador(request.user):
        ranking_servidores = Tarefa.objects.filter(
            tipo_fila=codigo_fila,
            siape_responsavel__isnull=False,
            siape_responsavel__siape__gt=0,  # Excluir SIAPE = 0 (tarefas sem responsável)
            ativa=True
        ).values(
            'siape_responsavel__siape',
            'siape_responsavel__nome_completo'
        ).annotate(
            total=Count('numero_protocolo_tarefa'),
            criticas=Count('numero_protocolo_tarefa', filter=Q(nivel_criticidade_calculado='CRÍTICA')),
            regulares=Count('numero_protocolo_tarefa', filter=Q(nivel_criticidade_calculado='REGULAR'))
        ).order_by('-total')[:20]

        # Calcular percentuais
        for servidor in ranking_servidores:
            if servidor['total'] > 0:
                servidor['percentual_criticas'] = round((servidor['criticas'] / servidor['total']) * 100, 1)
            else:
                servidor['percentual_criticas'] = 0

    context = {
        'info_fila': info_fila,
        'codigo_fila': codigo_fila,
        'total': total,
        'criticas': criticas,
        'regulares': regulares,
        'percentual_criticas': (criticas / total * 100) if total > 0 else 0,
        'status_counts': status_counts,
        'grafico_criticidade': grafico_criticidade,
        'grafico_status': grafico_status,
        'tarefas': tarefas_page,
        'servidores_fila': servidores_fila,
        'ranking_servidores': ranking_servidores,
        'filtro_criticidade': filtro_criticidade,
        'filtro_status': filtro_status,
        'filtro_servidor': filtro_servidor,
        'servidor_filtrado': servidor_filtrado,
        'ordenacao': ordenacao,
    }

    return render(request, 'tarefas/detalhe_fila.html', context)


# ============================================
# LISTA DE TAREFAS
# ============================================
@login_required
@user_passes_test(usuario_eh_coordenador)
def lista_tarefas(request):
    """
    Lista completa de tarefas com filtros OTIMIZADOS.

    Acesso: Apenas Coordenadores
    """

    tarefas = Tarefa.objects.select_related('siape_responsavel').filter(ativa=True)
    
    # Filtros
    protocolo = request.GET.get('protocolo', '').strip()
    if protocolo:
        tarefas = tarefas.filter(numero_protocolo_tarefa__icontains=protocolo)
    
    nivel = request.GET.get('nivel', '')
    if nivel:
        tarefas = tarefas.filter(nivel_criticidade_calculado=nivel)
    
    status = request.GET.get('status', '')
    if status:
        tarefas = tarefas.filter(status_tarefa=status)
    
    servidor = request.GET.get('servidor', '').strip()
    if servidor:
        tarefas = tarefas.filter(
            Q(siape_responsavel__siape__icontains=servidor) |
            Q(nome_profissional_responsavel__icontains=servidor)
        )
    
    servico = request.GET.get('servico', '').strip()
    if servico:
        tarefas = tarefas.filter(nome_servico__icontains=servico)
    
    data_inicio = request.GET.get('data_inicio', '')
    if data_inicio:
        tarefas = tarefas.filter(data_distribuicao_tarefa__gte=data_inicio)
    
    data_fim = request.GET.get('data_fim', '')
    if data_fim:
        tarefas = tarefas.filter(data_distribuicao_tarefa__lte=data_fim)
    
    # Ordenação (sem alteração, 'pontuacao_criticidade' agora é 100 ou 0)
    tarefas = tarefas.order_by('-pontuacao_criticidade', '-tempo_em_pendencia_em_dias')
    
    # Estatísticas (SIMPLIFICADO)
    stats = Tarefa.estatisticas_criticidade(tarefas)

    # Contagens adicionais para KPIs
    justificadas_count = tarefas.filter(tem_justificativa_ativa=True).count()
    com_ajuda_count = tarefas.filter(tem_solicitacao_ajuda=True).count()
    excluidas_count = tarefas.filter(servico_excluido_criticidade=True).count()

    kpis = {
        'total': stats['total'],
        'criticas': stats['criticas'],
        'regulares': stats['regulares'],
        'justificadas': justificadas_count,
        'com_ajuda': com_ajuda_count,
        'excluidas': excluidas_count,
    }
    
    # Paginação
    paginator = Paginator(tarefas, 50)
    page_number = request.GET.get('page')
    tarefas_paginadas = paginator.get_page(page_number)
    
    context = {
        'tarefas': tarefas_paginadas,
        'kpis': kpis,
        'filtros_ativos': {
            'protocolo': protocolo,
            'nivel': nivel,
            'status': status,
            'servidor': servidor,
            'servico': servico,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
        },
        # Níveis simplificados para o dropdown
        'niveis_disponiveis': ['CRÍTICA', 'REGULAR'],
        'status_disponiveis': Tarefa.objects.filter(ativa=True).values_list('status_tarefa', flat=True).distinct(),
    }
    
    return render(request, 'tarefas/lista_tarefas.html', context)


# ============================================
# LISTA DE SERVIDORES (COM OTIMIZAÇÃO N+1 APLICADA)
# ============================================
@login_required
@user_passes_test(usuario_eh_coordenador)
def lista_servidores(request):
    """
    Lista de servidores com resumo SIMPLIFICADO e OTIMIZADO.

    Acesso: Apenas Coordenadores
    """
    
    # Base de servidores que têm tarefas ativas
    servidores_com_tarefas = Tarefa.objects.filter(ativa=True).values('siape_responsavel').distinct().exclude(siape_responsavel__isnull=True)
    siapes = [s['siape_responsavel'] for s in servidores_com_tarefas]
    servidores = User.objects.filter(siape__in=siapes)
    
    # Filtros
    nome = request.GET.get('nome', '').strip()
    if nome:
        servidores = servidores.filter(nome_completo__icontains=nome)
    
    siape_filtro = request.GET.get('siape', '').strip()
    if siape_filtro:
        servidores = servidores.filter(siape__icontains=siape_filtro)
    
    gex = request.GET.get('gex', '').strip()
    if gex:
        servidores = servidores.filter(gex__icontains=gex)

    # Filtro de Revisão de Ofício
    tem_revisao_oficio = request.GET.get('tem_revisao_oficio', '')
    if tem_revisao_oficio == '1':
        servicos_revisao = [
            'Revisão de Ofício Identificada',
            'Revisão de Ofício Identificada - Benefício por Incapacidade',
            'Revisão de Ofício - Benefício por Incapacidade'
        ]
        # Filtrar apenas servidores que possuem tarefas de revisão de ofício
        siapes_com_revisao = Tarefa.objects.filter(
            nome_servico__in=servicos_revisao,
            ativa=True
        ).values_list('siape_responsavel__siape', flat=True).distinct()
        servidores = servidores.filter(siape__in=siapes_com_revisao)

    # --- INÍCIO DA OTIMIZAÇÃO ---
    # Montar lista com stats (OTIMIZADO COM ANNOTATE)
    # Isso faz UMA consulta ao banco, em vez de N+1
    servidores_annotated = servidores.annotate(
        total=Count('tarefas_sob_responsabilidade'),
        criticas=Count('tarefas_sob_responsabilidade', 
                       filter=Q(tarefas_sob_responsabilidade__nivel_criticidade_calculado='CRÍTICA')),
        regulares=Count('tarefas_sob_responsabilidade', 
                        filter=Q(tarefas_sob_responsabilidade__nivel_criticidade_calculado='REGULAR'))
    )

    # Ordenação (agora direto no QuerySet)
    ordem = request.GET.get('ordem', 'criticas')
    if ordem == 'criticas':
        servidores_ordenados = servidores_annotated.order_by('-criticas', 'nome_completo')
    elif ordem == 'total':
        servidores_ordenados = servidores_annotated.order_by('-total', 'nome_completo')
    elif ordem == 'nome':
        servidores_ordenados = servidores_annotated.order_by('nome_completo')
    else:
        # Fallback para a ordenação padrão
        servidores_ordenados = servidores_annotated.order_by('-criticas', 'nome_completo')

    # Paginação (usa o queryset ordenado)
    paginator = Paginator(servidores_ordenados, 20) 
    page_number = request.GET.get('page')
    servidores_paginados = paginator.get_page(page_number)
    
    # Stats gerais (calculados a partir do queryset anotado)
    total_servidores = servidores_annotated.count()
    total_criticas = sum(s.criticas for s in servidores_annotated) # Soma os valores já anotados
    
    context = {
        'servidores': servidores_paginados, # Envia o queryset paginado
        'filtros_ativos': {'nome': nome, 'siape': siape_filtro, 'gex': gex, 'ordem': ordem, 'tem_revisao_oficio': tem_revisao_oficio},
        'total_servidores': total_servidores,
        'total_criticas': total_criticas,
    }
    # --- FIM DA OTIMIZAÇÃO ---
    
    return render(request, 'tarefas/lista_servidores.html', context)


# ============================================
# DETALHES DO SERVIDOR
# ============================================
# ============================================
# DETALHES DO SERVIDOR
# ============================================
@login_required
def detalhe_servidor(request, siape):
    """
    Detalhes do servidor - VERSÃO COM FILAS v5.0

    Exibe cards das filas onde o servidor possui tarefas.
    Ao clicar em um card, vai para o detalhamento daquela fila específica.
    """
    from tarefas.filas import ORDEM_FILAS, obter_info_fila

    servidor = get_object_or_404(User, siape=siape)

    # Query otimizada: contar tarefas ativas por fila do servidor
    stats_por_fila = Tarefa.objects.filter(
        siape_responsavel=servidor,
        ativa=True
    ).values('tipo_fila').annotate(
        total=Count('numero_protocolo_tarefa'),
        criticas=Count('numero_protocolo_tarefa', filter=Q(nivel_criticidade_calculado='CRÍTICA')),
        regulares=Count('numero_protocolo_tarefa', filter=Q(nivel_criticidade_calculado='REGULAR'))
    )

    # Transformar em dicionário
    stats_dict = {item['tipo_fila']: item for item in stats_por_fila}

    # Construir lista ordenada de cards (apenas filas com tarefas do servidor)
    cards_filas = []
    for codigo_fila in ORDEM_FILAS:
        info = obter_info_fila(codigo_fila)
        stats = stats_dict.get(codigo_fila, {'total': 0, 'criticas': 0, 'regulares': 0})

        if stats['total'] > 0:  # Só mostrar filas onde o servidor tem tarefas
            cards_filas.append({
                'codigo': codigo_fila,
                'nome': info['nome'],
                'nome_completo': info['nome_completo'],
                'descricao': info['descricao'],
                'cor': info['cor'],
                'cor_bootstrap': info['cor_bootstrap'],
                'icone': info['icone'],
                'total': stats['total'],
                'criticas': stats['criticas'],
                'regulares': stats['regulares'],
                'percentual_criticas': (stats['criticas'] / stats['total'] * 100) if stats['total'] > 0 else 0,
            })

    # Totais gerais do servidor
    total_geral = sum(c['total'] for c in cards_filas)
    criticas_geral = sum(c['criticas'] for c in cards_filas)
    regulares_geral = sum(c['regulares'] for c in cards_filas)

    context = {
        'servidor': servidor,
        'cards_filas': cards_filas,
        'total_geral': total_geral,
        'criticas_geral': criticas_geral,
        'regulares_geral': regulares_geral,
        'percentual_criticas_geral': (criticas_geral / total_geral * 100) if total_geral > 0 else 0,
        'eh_proprio_servidor': request.user == servidor,
    }

    return render(request, 'tarefas/detalhe_servidor.html', context)


# ============================================
# DETALHES DA TAREFA (REDESENHADO)
# ============================================
@login_required
def detalhe_tarefa(request, protocolo):
    """
    Detalhes completos da tarefa - v4.0
    """

    tarefa = get_object_or_404(Tarefa, numero_protocolo_tarefa=protocolo, ativa=True)
    
    # Info básica (sem alteração)
    info_basica = {
        'protocolo': tarefa.numero_protocolo_tarefa,
        'servico': tarefa.nome_servico,
        'status': tarefa.status_tarefa,
        'servidor': tarefa.nome_profissional_responsavel or 'Não atribuído',
        'siape': tarefa.siape_responsavel.siape if tarefa.siape_responsavel else 'N/A',
        'gex': tarefa.nome_gex_responsavel or 'N/A',
    }
    
    # Datas (sem alteração)
    datas = {
        'distribuicao': tarefa.data_distribuicao_tarefa,
        'ultima_atualizacao': tarefa.data_ultima_atualizacao,
        'prazo': tarefa.data_prazo,
        'inicio_exigencia': tarefa.data_inicio_ultima_exigencia,
        'fim_exigencia': tarefa.data_fim_ultima_exigencia,
        'processamento': tarefa.data_processamento_tarefa,
        'calculo_criticidade': tarefa.data_calculo_criticidade,
    }
    
    # Tempos (sem alteração)
    tempos = {
        'pendencia': tarefa.tempo_em_pendencia_em_dias,
        'exigencia': tarefa.tempo_em_exigencia_em_dias,
        'ultima_exigencia': tarefa.tempo_ultima_exigencia_em_dias,
        'ate_distribuicao': tarefa.tempo_ate_ultima_distribuicao_tarefa_em_dias,
        'com_servidor': tarefa.dias_com_servidor,
    }
    
    # Criticidade (atualizado para incluir nome amigável da regra)
    criticidade = {
        'nivel': tarefa.nivel_criticidade_calculado,
        'regra': tarefa.regra_aplicada_nome,  # ← ATUALIZADO: Nome amigável da regra
        'regra_codigo': tarefa.regra_aplicada_calculado,  # ← NOVO: Código técnico da regra
        'alerta': tarefa.alerta_criticidade_calculado,
        'descricao': tarefa.descricao_criticidade_calculado,
        'dias_pendente': tarefa.dias_pendente_criticidade_calculado,
        'prazo_limite': tarefa.prazo_limite_criticidade_calculado,
        'cor': tarefa.cor_criticidade_calculado,
        'emoji': tarefa.emoji_criticidade,
        'pontuacao': tarefa.pontuacao_criticidade,
    }
    
    # Info da regra (usar código da regra para buscar informações)
    regras_info = get_regras_info()
    regra_info = regras_info.get(criticidade['regra_codigo'], {
        'nome': 'Sem classificação',
        'descricao': 'Não se enquadra em nenhuma regra.',
        'prazo': 'N/A',
        'condicoes': [],
        'calculo': 'N/A',
        'classificacao': [],
        'badge_class': 'bg-secondary',
        'icon': 'fas fa-question-circle',
    })
    
    # Adicionar detalhes específicos da regra (CLASSIFICAÇÃO ATUALIZADA)
    if criticidade['regra_codigo'] == 'REGRA_1_EXIGENCIA_CUMPRIDA':
        regra_info['condicoes'] = [
            'Status = "Pendente"',
            'Descrição = "Exigência cumprida"',
            'Servidor cadastrou a exigência',
            'Exigência foi cumprida',
        ]
        regra_info['calculo'] = f'dias_desde_cumprimento = HOJE - {tarefa.data_fim_ultima_exigencia.strftime("%d/%m/%Y") if tarefa.data_fim_ultima_exigencia else "N/A"}'
        regra_info['classificacao'] = ['≤ 7 dias = REGULAR', '> 7 dias = CRÍTICA']

    elif criticidade['regra_codigo'] == 'REGRA_2_CUMPRIMENTO_EXIGENCIA':
        regra_info['condicoes'] = [
            'Status = "Cumprimento de exigência"',
            'Descrição = "Em cumprimento"',
            f'Data prazo: {tarefa.data_prazo.strftime("%d/%m/%Y") if tarefa.data_prazo else "N/A"}',
        ]
        regra_info['calculo'] = f'dias_apos_prazo = HOJE - ({tarefa.data_prazo.strftime("%d/%m/%Y") if tarefa.data_prazo else "N/A"} + 5 dias)'
        regra_info['classificacao'] = ['Dentro do prazo = REGULAR', 'Prazo vencido = CRÍTICA']

    elif criticidade['regra_codigo'] == 'REGRA_3_PRIMEIRA_ACAO_SEM_EXIGENCIA':
        regra_info['condicoes'] = [
            'Status = "Pendente"',
            'Nunca entrou em exigência',
            'Sem data de início',
        ]
        regra_info['calculo'] = f'dias_com_servidor = {tarefa.tempo_em_pendencia_em_dias} - {tarefa.tempo_ate_ultima_distribuicao_tarefa_em_dias}'
        regra_info['classificacao'] = ['≤ 10 dias = REGULAR', '> 10 dias = CRÍTICA']

    elif criticidade['regra_codigo'] == 'REGRA_4_PRIMEIRA_ACAO_COM_EXIGENCIA':
        regra_info['condicoes'] = [
            'Status = "Pendente"',
            'Exigência não cumprida',
            'Data início exigência existe',
        ]
        regra_info['calculo'] = f'dias_com_servidor = {tarefa.tempo_em_pendencia_em_dias} - {tarefa.tempo_ate_ultima_distribuicao_tarefa_em_dias}'
        regra_info['classificacao'] = ['≤ 10 dias = REGULAR', '> 10 dias = CRÍTICA']
    
    context = {
        'tarefa': tarefa,
        'info_basica': info_basica,
        'datas': datas,
        'tempos': tempos,
        'criticidade': criticidade,
        'regra_info': regra_info,
    }
    
    return render(request, 'tarefas/detalhe_tarefa.html', context)


# ============================================
# DASHBOARD SERVIDOR (REDIRECIONA)
# ============================================
@login_required
def dashboard_servidor(request):
    """Redireciona para detalhes do servidor."""
    return redirect('tarefas:detalhe_servidor', siape=request.user.siape)


# ============================================
# API JSON
# ============================================
@login_required
def api_estatisticas_json(request):
    """Retorna estatísticas em JSON."""
    tarefas = Tarefa.objects.filter(ativa=True)
    stats = Tarefa.estatisticas_criticidade(tarefas) # Retorna o novo formato
    return JsonResponse(stats)


# ============================================
# CONFIGURAÇÕES DO SISTEMA
# ============================================
@login_required
@user_passes_test(usuario_eh_coordenador)
def configuracoes(request):
    """Página de configurações do sistema (apenas coordenadores)."""
    context = {
        'params': ParametrosAnalise.get_configuracao_ativa(),
        'total_tarefas': Tarefa.objects.filter(ativa=True).count(),
    }
    return render(request, 'tarefas/configuracoes.html', context)


@login_required
@user_passes_test(usuario_eh_coordenador)
def recalcular_criticidades(request):
    """
    Recalcula a criticidade de todas as tarefas do sistema.
    Apenas coordenadores podem executar esta ação.
    """
    from django.contrib import messages
    from django.utils import timezone
    from tarefas.analisador import obter_analisador

    if request.method == 'POST':
        try:
            analisador = obter_analisador()
            tarefas = Tarefa.objects.filter(ativa=True)
            total = tarefas.count()

            # Contadores
            atualizadas = 0
            erros = 0

            # Processar em lote
            for tarefa in tarefas:
                try:
                    # Atualizar flags de justificativa
                    if hasattr(tarefa, 'atualizar_flags_justificativa'):
                        tarefa.atualizar_flags_justificativa()

                    # Analisar criticidade
                    resultado = analisador.analisar_tarefa(tarefa)

                    # Atualizar campos
                    tarefa.nivel_criticidade_calculado = resultado['nivel']
                    tarefa.regra_aplicada_calculado = resultado['regra']
                    tarefa.alerta_criticidade_calculado = resultado['alerta']
                    tarefa.descricao_criticidade_calculado = resultado['descricao']
                    tarefa.dias_pendente_criticidade_calculado = resultado['dias_pendente']
                    tarefa.prazo_limite_criticidade_calculado = resultado['prazo_limite']
                    tarefa.cor_criticidade_calculado = resultado['cor']

                    # Calcular pontuação
                    ordem_severidade = {
                        'CRÍTICA': 5,
                        'JUSTIFICADA': 4,
                        'EXCLUÍDA': 3,
                        'REGULAR': 2,
                    }
                    tarefa.pontuacao_criticidade = ordem_severidade.get(resultado['nivel'], 0)
                    tarefa.data_calculo_criticidade = timezone.now()

                    # Salvar
                    tarefa.save(update_fields=[
                        'nivel_criticidade_calculado',
                        'regra_aplicada_calculado',
                        'alerta_criticidade_calculado',
                        'descricao_criticidade_calculado',
                        'dias_pendente_criticidade_calculado',
                        'prazo_limite_criticidade_calculado',
                        'pontuacao_criticidade',
                        'cor_criticidade_calculado',
                        'data_calculo_criticidade',
                        'tem_justificativa_ativa',
                        'tem_solicitacao_ajuda',
                        'servico_excluido_criticidade',
                    ])

                    atualizadas += 1

                except Exception as e:
                    erros += 1
                    print(f"Erro ao recalcular tarefa {tarefa.numero_protocolo_tarefa}: {e}")

            # Buscar estatísticas atualizadas
            stats = Tarefa.estatisticas_criticidade()

            # Mensagem de sucesso
            messages.success(
                request,
                f'Recálculo concluído com sucesso! '
                f'{atualizadas} tarefas atualizadas. '
                f'Críticas: {stats.get("criticas", 0)} | '
                f'Regulares: {stats.get("regulares", 0)}'
            )

            if erros > 0:
                messages.warning(request, f'{erros} tarefas com erro durante o recálculo.')

        except Exception as e:
            messages.error(request, f'Erro ao recalcular criticidades: {str(e)}')

    return redirect('tarefas:configuracoes')

# ============================================
# VIEWS PARA AÇÕES AUTOMATIZADAS
# ============================================

from django.contrib import messages
from django.views.decorators.http import require_POST
from tarefas.services.acoes_service import AcoesService
from tarefas.services.email_service import EmailService
from tarefas.models import BloqueioServidor, SolicitacaoNotificacao, TemplateEmail

@login_required
@user_passes_test(usuario_eh_coordenador)
@require_POST
def solicitar_bloqueio_servidor(request, siape):
    """Solicita bloqueio de caixa para um servidor."""
    codigo_fila = request.POST.get('codigo_fila')
    observacoes = request.POST.get('observacoes', '')

    try:
        servidor = get_object_or_404(User, siape=siape)

        # Verificar se já existe bloqueio ativo
        if BloqueioServidor.servidor_esta_bloqueado(siape, codigo_fila):
            messages.warning(request, f'O servidor {siape} já está bloqueado para a fila {codigo_fila}.')
            return redirect(request.META.get('HTTP_REFERER', '/'))

        # Criar solicitação
        bloqueio = AcoesService.solicitar_bloqueio(
            servidor=servidor,
            codigo_fila=codigo_fila,
            solicitado_por=request.user,
            observacoes=observacoes
        )

        messages.success(
            request,
            f'Solicitação de bloqueio criada com sucesso! '
            f'ID: {bloqueio.id} | Servidor: {siape} | Fila: {codigo_fila}'
        )

    except Exception as e:
        messages.error(request, f'Erro ao solicitar bloqueio: {str(e)}')

    return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
@user_passes_test(usuario_eh_coordenador)
@require_POST
def solicitar_desbloqueio_servidor(request, siape):
    """Solicita desbloqueio de caixa para um servidor."""
    codigo_fila = request.POST.get('codigo_fila')
    observacoes = request.POST.get('observacoes', '')

    try:
        servidor = get_object_or_404(User, siape=siape)

        # Criar solicitação
        desbloqueio = AcoesService.solicitar_desbloqueio(
            servidor=servidor,
            codigo_fila=codigo_fila,
            solicitado_por=request.user,
            observacoes=observacoes
        )

        messages.success(
            request,
            f'Solicitação de desbloqueio criada com sucesso! '
            f'ID: {desbloqueio.id} | Servidor: {siape} | Fila: {codigo_fila}'
        )

    except Exception as e:
        messages.error(request, f'Erro ao solicitar desbloqueio: {str(e)}')

    return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
@user_passes_test(usuario_eh_coordenador)
@require_POST
def solicitar_notificacao_pgb(request, siape):
    """Solicita criação de tarefa PGB (notificação)."""
    tipo_notificacao = request.POST.get('tipo_notificacao')
    observacoes = request.POST.get('observacoes', '')

    try:
        servidor = get_object_or_404(User, siape=siape)

        # Validar tipo
        if tipo_notificacao not in [SolicitacaoNotificacao.TIPO_PRIMEIRA, SolicitacaoNotificacao.TIPO_SEGUNDA]:
            messages.error(request, 'Tipo de notificação inválido.')
            return redirect(request.META.get('HTTP_REFERER', '/'))

        # Criar solicitação
        notificacao = AcoesService.solicitar_notificacao_pgb(
            servidor=servidor,
            tipo_notificacao=tipo_notificacao,
            solicitado_por=request.user,
            observacoes=observacoes
        )

        tipo_display = notificacao.get_tipo_notificacao_display()
        messages.success(
            request,
            f'Solicitação criada com sucesso! '
            f'ID: {notificacao.id} | Tipo: {tipo_display} | Servidor: {siape}'
        )

    except Exception as e:
        messages.error(request, f'Erro ao solicitar notificação: {str(e)}')

    return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
@user_passes_test(usuario_eh_coordenador)
@require_POST
def enviar_email_servidor(request, siape):
    """Envia email para um servidor usando template."""
    template_nome = request.POST.get('template_nome')

    try:
        servidor = get_object_or_404(User, siape=siape)

        if not servidor.email:
            messages.error(request, f'Servidor {siape} não possui email cadastrado.')
            return redirect(request.META.get('HTTP_REFERER', '/'))

        # Buscar tarefas do servidor para incluir no email
        tarefas = Tarefa.objects.filter(
            siape_responsavel=servidor,
            ativa=True
        ).order_by('-pontuacao_criticidade')[:20]

        # Preparar contexto adicional
        contexto_extra = {
            'total_tarefas': tarefas.count(),
            'tarefas_criticas': tarefas.filter(nivel_criticidade_calculado='CRÍTICA').count(),
            'lista_tarefas': EmailService().criar_lista_tarefas_html(tarefas),
        }

        # Enviar email
        email_service = EmailService()
        sucesso, historico = email_service.enviar_email_por_template(
            template_nome=template_nome,
            servidor=servidor,
            enviado_por=request.user,
            contexto_extra=contexto_extra
        )

        if sucesso:
            messages.success(
                request,
                f'Email enviado com sucesso para {servidor.email}! '
                f'ID do histórico: {historico.id}'
            )
        else:
            messages.error(
                request,
                f'Erro ao enviar email: {historico.mensagem_erro}'
            )

    except TemplateEmail.DoesNotExist:
        messages.error(request, f'Template "{template_nome}" não encontrado.')
    except Exception as e:
        messages.error(request, f'Erro ao enviar email: {str(e)}')

    return redirect(request.META.get('HTTP_REFERER', '/'))


@login_required
def verificar_status_servidor(request, siape):
    """Retorna status do servidor (bloqueios, notificações, etc) em JSON."""
    codigo_fila = request.GET.get('codigo_fila', '')

    try:
        # Verificar bloqueio (se codigo_fila específico foi fornecido)
        if codigo_fila:
            status_bloqueio = AcoesService.verificar_status_servidor(siape, codigo_fila)
            bloqueado = status_bloqueio['bloqueado']
            tem_solicitacao_pendente = status_bloqueio['tem_solicitacao_pendente']
        else:
            # Sem fila específica, retornar False (usado em detalhe_servidor)
            bloqueado = False
            tem_solicitacao_pendente = False

        # Listar todas as filas bloqueadas (para detalhe_servidor)
        from tarefas.filas import ORDEM_FILAS
        filas_bloqueadas = []
        for fila_cod in ORDEM_FILAS:
            if BloqueioServidor.servidor_esta_bloqueado(siape, fila_cod):
                filas_bloqueadas.append(fila_cod)

        # Verificar se há solicitações pendentes (qualquer fila)
        tem_solicitacao_pendente_geral = BloqueioServidor.objects.filter(
            servidor__siape=siape,
            status__in=[BloqueioServidor.STATUS_PENDENTE, BloqueioServidor.STATUS_PROCESSANDO]
        ).exists()

        # Verificar notificações
        tem_notificacao = SolicitacaoNotificacao.servidor_tem_notificacao_ativa(siape)

        # Verificar tarefas de revisão de ofício
        servicos_revisao = [
            'Revisão de Ofício Identificada',
            'Revisão de Ofício Identificada - Benefício por Incapacidade',
            'Revisão de Ofício - Benefício por Incapacidade'
        ]
        tem_revisao_oficio = Tarefa.objects.filter(
            siape_responsavel__siape=siape,
            nome_servico__in=servicos_revisao,
            ativa=True
        ).exists()

        total_revisao_oficio = Tarefa.objects.filter(
            siape_responsavel__siape=siape,
            nome_servico__in=servicos_revisao,
            ativa=True
        ).count()

        # Contar total de tarefas na fila específica (se fornecida)
        total_tarefas = 0
        if codigo_fila:
            total_tarefas = Tarefa.objects.filter(
                siape_responsavel__siape=siape,
                tipo_fila=codigo_fila,
                ativa=True
            ).count()

        return JsonResponse({
            'success': True,
            'bloqueado': bloqueado,  # Para fila específica
            'filas_bloqueadas': filas_bloqueadas,  # Lista de todas as filas bloqueadas
            'tem_solicitacao_pendente': tem_solicitacao_pendente or tem_solicitacao_pendente_geral,
            'tem_notificacao': tem_notificacao,
            'tem_revisao_oficio': tem_revisao_oficio,
            'total_revisao_oficio': total_revisao_oficio,
            'total_tarefas': total_tarefas,  # Total de tarefas na fila (para arquivo em lote)
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


# ============================================
# EXPORTAÇÃO DE RELATÓRIOS
# ============================================
@login_required
def exportar_tarefas_fila_servidor_excel(request, codigo_fila):
    """
    Exporta relatório em Excel com as tarefas de um servidor em uma fila específica.

    Parâmetros:
    - codigo_fila: Código da fila (ex: PGB, REVISAO, etc)
    - servidor: SIAPE do servidor (via GET parameter)

    Retorna arquivo Excel com informações detalhadas das tarefas.
    """
    from tarefas.filas import obter_info_fila

    # Obter SIAPE do servidor do parâmetro GET
    siape_servidor = request.GET.get('servidor')

    if not siape_servidor:
        return HttpResponse('SIAPE do servidor não informado', status=400)

    # Buscar informações do servidor
    try:
        servidor = User.objects.get(siape=siape_servidor)
    except User.DoesNotExist:
        return HttpResponse('Servidor não encontrado', status=404)

    # Obter informações da fila
    info_fila = obter_info_fila(codigo_fila)

    # Buscar tarefas do servidor nesta fila
    tarefas = Tarefa.objects.filter(
        tipo_fila=codigo_fila,
        siape_responsavel__siape=siape_servidor,
        siape_responsavel__isnull=False,
        ativa=True
    ).select_related('siape_responsavel').order_by('-pontuacao_criticidade', 'data_distribuicao_tarefa')

    # Verificar se há tarefas
    if not tarefas.exists():
        return HttpResponse('Nenhuma tarefa encontrada para este servidor nesta fila', status=404)

    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Tarefas {codigo_fila}"

    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    wrap_alignment = Alignment(wrap_text=True, vertical='top')

    # Título do relatório
    ws.merge_cells('A1:M1')
    title_cell = ws['A1']
    title_cell.value = f"RELATÓRIO DE TAREFAS - {info_fila['nome_completo']}"
    title_cell.font = Font(bold=True, size=14, color="366092")
    title_cell.alignment = center_alignment

    # Informações do servidor
    ws.merge_cells('A2:M2')
    info_cell = ws['A2']
    info_cell.value = f"Servidor: {servidor.nome_completo} (SIAPE: {servidor.siape})"
    info_cell.font = Font(bold=True, size=11)
    info_cell.alignment = center_alignment

    # Data de geração
    ws.merge_cells('A3:M3')
    date_cell = ws['A3']
    date_cell.value = f"Data de Geração: {date.today().strftime('%d/%m/%Y')}"
    date_cell.alignment = center_alignment

    # Linha em branco
    ws.append([])

    # Cabeçalhos
    headers = [
        'Protocolo',
        'Serviço',
        'Interessado',
        'Status',
        'Data Entrada',
        'Dias Parado',
        'Prazo (dias)',
        'Criticidade',
        'Pontuação',
        'Unidade',
        'Última Movimentação',
        'Responsável',
        'Observações'
    ]

    ws.append(headers)
    header_row = ws[5]

    for cell in header_row:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border

    # Definir cores por criticidade
    criticidade_colors = {
        'CRÍTICA': PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid"),
        'REGULAR': PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid"),
    }

    # Adicionar dados das tarefas
    for tarefa in tarefas:
        row_data = [
            tarefa.numero_protocolo_tarefa or '',
            tarefa.nome_servico or '',
            '',  # Interessado (campo removido)
            tarefa.status_tarefa or '',
            tarefa.data_distribuicao_tarefa.strftime('%d/%m/%Y') if tarefa.data_distribuicao_tarefa else '',
            tarefa.tempo_em_pendencia_em_dias or 0,
            tarefa.prazo_limite_criticidade_calculado or 0,
            tarefa.nivel_criticidade_calculado or '',
            tarefa.pontuacao_criticidade or 0,
            tarefa.codigo_unidade_tarefa or '',
            tarefa.data_ultima_atualizacao.strftime('%d/%m/%Y') if tarefa.data_ultima_atualizacao else '',
            f"{tarefa.siape_responsavel.nome_completo} ({tarefa.siape_responsavel.siape})" if tarefa.siape_responsavel else '',
            ''  # Observações (campo removido)
        ]
        ws.append(row_data)

        # Aplicar estilos na linha
        current_row = ws.max_row
        criticidade = tarefa.nivel_criticidade_calculado or 'REGULAR'
        fill = criticidade_colors.get(criticidade, PatternFill())

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.border = border
            cell.fill = fill

            # Alinhamento específico por coluna
            if col_num in [1, 4, 5, 6, 7, 8, 9, 10, 11]:  # Colunas centralizadas
                cell.alignment = center_alignment
            else:
                cell.alignment = wrap_alignment

    # Ajustar largura das colunas
    column_widths = {
        'A': 18,  # Protocolo
        'B': 40,  # Serviço
        'C': 30,  # Interessado
        'D': 20,  # Status
        'E': 12,  # Data Entrada
        'F': 12,  # Dias Parado
        'G': 12,  # Prazo
        'H': 12,  # Criticidade
        'I': 12,  # Pontuação
        'J': 12,  # Unidade
        'K': 15,  # Última Movimentação
        'L': 30,  # Responsável
        'M': 40,  # Observações
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Congelar painéis (cabeçalho)
    ws.freeze_panes = 'A6'

    # Adicionar rodapé com estatísticas
    ws.append([])
    stats_row = ws.max_row + 1

    total_tarefas = tarefas.count()
    criticas = tarefas.filter(nivel_criticidade_calculado='CRÍTICA').count()
    regulares = tarefas.filter(nivel_criticidade_calculado='REGULAR').count()

    ws.merge_cells(f'A{stats_row}:M{stats_row}')
    stats_cell = ws[f'A{stats_row}']
    stats_cell.value = f"TOTAL DE TAREFAS: {total_tarefas} | CRÍTICAS: {criticas} | REGULARES: {regulares}"
    stats_cell.font = Font(bold=True, size=11)
    stats_cell.alignment = center_alignment
    stats_cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    # Preparar response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Tarefas_{codigo_fila}_{servidor.siape}_{date.today().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Salvar workbook no response
    wb.save(response)

    return response


# ============================================
# AÇÕES EM LOTE - GERAÇÃO DE ARQUIVO CSV
# ============================================

@login_required
@user_passes_test(usuario_eh_coordenador)
def obter_servicos_servidor_fila(request, siape, codigo_fila):
    """
    Endpoint API para obter lista de serviços distintos
    que um servidor possui em uma fila específica.
    """
    try:
        servidor = get_object_or_404(User, siape=siape)

        # Buscar serviços distintos
        servicos = Tarefa.objects.filter(
            siape_responsavel=servidor,
            tipo_fila=codigo_fila,
            ativa=True
        ).values_list('nome_servico', flat=True).distinct().order_by('nome_servico')

        return JsonResponse({
            'success': True,
            'servicos': list(servicos)
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@user_passes_test(usuario_eh_coordenador)
def gerar_arquivo_acao_lote(request, siape, codigo_fila):
    """
    Gera arquivo CSV para ação em lote (remover responsável ou transferir tarefas).

    Formatos suportados:
    - Tipo 14: Remover Responsável em Lote
    - Tipo 12: Transferir Tarefa em Lote
    """
    from tarefas.models import HistoricoAcaoLote
    from datetime import datetime
    import csv
    from io import StringIO

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Método não permitido'}, status=405)

    try:
        servidor = get_object_or_404(User, siape=siape)

        # Obter parâmetros do formulário
        tipo_acao = request.POST.get('tipo_acao')  # REMOVER ou TRANSFERIR
        criterio_selecao = request.POST.get('criterio_selecao')  # TODAS, SERVICO, QUANTIDADE, MANUAL

        # Validar tipo de ação
        if tipo_acao not in ['REMOVER', 'TRANSFERIR']:
            return JsonResponse({'success': False, 'error': 'Tipo de ação inválido'}, status=400)

        # Construir queryset base
        tarefas_query = Tarefa.objects.filter(
            siape_responsavel=servidor,
            tipo_fila=codigo_fila,
            ativa=True
        )

        # Aplicar critério de seleção
        servico_selecionado = ''
        protocolos_manuais = []

        if criterio_selecao == 'TODAS':
            # Ordenar por dias pendentes (mais dias primeiro)
            tarefas = tarefas_query.order_by('-dias_pendente_criticidade_calculado')

        elif criterio_selecao == 'SERVICO':
            servico_selecionado = request.POST.get('servico_nome', '')
            if not servico_selecionado:
                return JsonResponse({'success': False, 'error': 'Serviço não informado'}, status=400)

            tarefas = tarefas_query.filter(
                nome_servico=servico_selecionado
            ).order_by('-dias_pendente_criticidade_calculado')

        elif criterio_selecao == 'QUANTIDADE':
            quantidade = request.POST.get('quantidade', 0)
            try:
                quantidade = int(quantidade)
                if quantidade <= 0:
                    return JsonResponse({'success': False, 'error': 'Quantidade inválida'}, status=400)
            except ValueError:
                return JsonResponse({'success': False, 'error': 'Quantidade deve ser um número'}, status=400)

            # Ordenar por dias pendentes e pegar N primeiras
            tarefas = tarefas_query.order_by('-dias_pendente_criticidade_calculado')[:quantidade]

        elif criterio_selecao == 'MANUAL':
            protocolos_texto = request.POST.get('protocolos_manuais', '')
            if not protocolos_texto:
                return JsonResponse({'success': False, 'error': 'Nenhum protocolo informado'}, status=400)

            # Processar lista de protocolos (separados por vírgula, espaço ou quebra de linha)
            import re
            protocolos_manuais = re.split(r'[,\s\n]+', protocolos_texto.strip())
            protocolos_manuais = [p.strip() for p in protocolos_manuais if p.strip()]

            if not protocolos_manuais:
                return JsonResponse({'success': False, 'error': 'Nenhum protocolo válido informado'}, status=400)

            tarefas = tarefas_query.filter(numero_protocolo_tarefa__in=protocolos_manuais)

        else:
            return JsonResponse({'success': False, 'error': 'Critério de seleção inválido'}, status=400)

        # Verificar se há tarefas
        if not tarefas.exists():
            return JsonResponse({'success': False, 'error': 'Nenhuma tarefa encontrada com os critérios informados'}, status=400)

        # Parâmetros específicos para transferência
        uo_destino = ''
        despacho = ''

        if tipo_acao == 'TRANSFERIR':
            uo_destino = request.POST.get('uo_destino', '').strip()
            despacho = request.POST.get('despacho', '').strip()

            if not uo_destino:
                return JsonResponse({'success': False, 'error': 'UO de destino é obrigatória para transferência'}, status=400)

            if not despacho or len(despacho) < 30:
                return JsonResponse({'success': False, 'error': 'Despacho deve ter no mínimo 30 caracteres'}, status=400)

        # Gerar arquivo CSV
        output = StringIO()

        if tipo_acao == 'REMOVER':
            # Tipo 14: Remover Responsável em Lote
            output.write('14\n')
            output.write('"Protocolo da Tarefa"\n')

            for tarefa in tarefas:
                output.write(f'"{tarefa.numero_protocolo_tarefa}"\n')

            tipo_arquivo = 'remover'
            tipo_codigo = '14'

        else:  # TRANSFERIR
            # Tipo 12: Transferir Tarefa em Lote
            output.write('12\n')
            output.write('"Protocolo da Tarefa;UO de Destino da Tarefa;Despacho"\n')

            for tarefa in tarefas:
                # Escapar aspas duplas no despacho (duplicando-as)
                despacho_escapado = despacho.replace('"', '""')
                output.write(f'"{tarefa.numero_protocolo_tarefa};{uo_destino};{despacho_escapado}"\n')

            tipo_arquivo = 'transferir'
            tipo_codigo = '12'

        # Registrar no histórico
        protocolos_lista = list(tarefas.values_list('numero_protocolo_tarefa', flat=True))

        nome_arquivo = f'acao_lote_{tipo_arquivo}_{siape}_{codigo_fila}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

        historico = HistoricoAcaoLote.objects.create(
            servidor=servidor,
            codigo_fila=codigo_fila,
            tipo_acao='REMOVER_RESPONSAVEL' if tipo_acao == 'REMOVER' else 'TRANSFERIR_TAREFA',
            criterio_selecao=criterio_selecao,
            servico_selecionado=servico_selecionado,
            quantidade_tarefas=len(protocolos_lista),
            uo_destino=uo_destino,
            despacho=despacho,
            gerado_por=request.user,
            nome_arquivo=nome_arquivo,
            protocolos_incluidos=','.join(protocolos_lista)
        )

        # Preparar response
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'

        # Escrever conteúdo com encoding UTF-8 (padrão CSV brasileiro)
        response.write(output.getvalue())

        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({
            'success': False,
            'error': f'Erro ao gerar arquivo: {str(e)}'
        }, status=500)
